"""One-off export: HP 500 (Sprocket Panorama) ATP workbook -> JSON + CSV."""
import csv
import json
from pathlib import Path

SRC = Path(r"D:\ATP Sheets\HP 500 (Sprocket Panorama) ATP (4).xlsx")
OUT_DIR = Path(__file__).resolve().parent


def sheet_code(name: str) -> str:
    return (
        name.replace(" ", "_")
        .replace("+", "plus")
        .replace("/", "_")
        .replace("&", "and")
    )


def main() -> None:
    import openpyxl

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.load_workbook(SRC, data_only=True)
    cases: list[dict] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        code = sheet_code(sheet_name)
        for r in range(2, ws.max_row + 1):
            row = {
                (headers[i - 1] if headers[i - 1] else f"col{i}"): ws.cell(r, i).value
                for i in range(1, ws.max_column + 1)
            }
            if not any(v is not None and str(v).strip() for v in row.values()):
                continue
            entry: dict = {
                "atp_id": f"{code}_{r:03d}",
                "sheet": sheet_name,
                "excel_row": r,
                "test_objective": row.get("Test Objective"),
                "steps": row.get("Steps"),
                "test_data": row.get("Test data"),
                "expected_result": row.get("Expected result"),
                "pass_fail": row.get("PASS/FAIL"),
                "comments": row.get("Comments"),
            }
            # SL sheet: human "Test N" = Excel row (N+1) → atp_id SL_{row:03d}
            # e.g. Test 1 → row 2 → SL_002, Test 2 → row 3 → SL_003
            if sheet_name == "SL":
                entry["sl_test_number"] = r - 1
                entry["sl_atp_id"] = f"SL_{r:03d}"
            cases.append(entry)

    json_path = OUT_DIR / "hp500_panorama_atp_cases.json"
    json_path.write_text(
        json.dumps(cases, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )

    csv_path = OUT_DIR / "hp500_panorama_atp_cases.csv"
    fieldnames = [
        "atp_id",
        "sheet",
        "excel_row",
        "sl_test_number",
        "sl_atp_id",
        "test_objective",
        "test_data",
        "steps",
        "expected_result",
        "pass_fail",
        "comments",
    ]
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        w.writeheader()
        for c in cases:
            flat = {}
            for k in fieldnames:
                v = c.get(k)
                if v is None:
                    flat[k] = ""
                elif isinstance(v, str):
                    flat[k] = v.replace("\r\n", "\n").replace("\n", " | ")
                else:
                    flat[k] = v
            w.writerow(flat)

    print(f"cases={len(cases)}")
    print(f"json={json_path}")
    print(f"csv={csv_path}")


if __name__ == "__main__":
    main()
